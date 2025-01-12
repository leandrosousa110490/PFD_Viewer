import sys
import os
import shutil
import tempfile
import json
from collections import deque

# PDF Handling
import fitz  # PyMuPDF

# Excel exports only (remove Word)
import openpyxl
import pandas as pd
import tempfile

# PyQt5 imports
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QFileDialog,
    QAction, QToolBar, QMessageBox, QLabel, QVBoxLayout,
    QScrollArea, QWidget, QLineEdit, QDialog, QHBoxLayout,
    QSpinBox, QComboBox, QPushButton, QStackedWidget,
    QColorDialog, QTextEdit, QMenu, QMenuBar,
    QTableWidget, QTableWidgetItem, QTextBrowser
)
from PyQt5.QtGui import (
    QCursor, QFont, QColor, QPixmap, QImage, QTextCursor
)
from PyQt5.QtCore import Qt


###############################################################################
# PageWidget: Container for each PDF page’s rendered image
###############################################################################
class PageWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(self.layout)

        # Create label for PDF page
        self.page_label = QLabel()
        self.page_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.page_label)

        # Context menu
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.page_label.setContextMenuPolicy(Qt.CustomContextMenu)
        self.page_label.customContextMenuRequested.connect(self.show_context_menu)

    def show_context_menu(self, position):
        menu = QMenu(self)
        
        # Get parent PDFViewWidget
        parent = self.parent()
        while parent and not isinstance(parent, PDFViewWidget):
            parent = parent.parent()
            
        if parent and isinstance(parent, PDFViewWidget):
            # Add Edit/Stop Editing toggle
            if not parent.edit_mode:
                edit_action = menu.addAction("Edit")
                edit_action.triggered.connect(parent.start_edit_mode)
            else:
                stop_edit_action = menu.addAction("Stop Editing")
                stop_edit_action.triggered.connect(parent.stop_edit_mode)
            
            menu.addSeparator()
            close_page_action = menu.addAction("Close This Page")
            try:
                page_index = parent.page_widgets.index(self)
                close_page_action.triggered.connect(
                    lambda: parent.remove_page(page_index)
                )
            except ValueError:
                pass

        menu.exec_(self.mapToGlobal(position))


###############################################################################
# PDFViewWidget: Displays a single PDF (multiple pages) & supports editing
###############################################################################
class PDFViewWidget(QWidget):
    """
    Displays one PDF file in multiple pages (PageWidgets).
    Allows in-place text editing where the user clicks.
    """
    def __init__(self, pdf_file_path, parent=None):
        super().__init__(parent)

        self.original_pdf_path = pdf_file_path
        self.doc = None
        self.zoom = 1.0
        self.edit_mode = False
        self.text_placement_mode = False
        self.current_page_index = None
        self.undo_stack = deque(maxlen=50)  # Store last 50 actions
        self.redo_stack = deque(maxlen=50)

        # Map from GUI font names to built-in PDF fonts (to fix "need font file" error).
        self.font_map = {
            "Arial": "helv",              # "Helvetica" base-14
            "Times New Roman": "Times-Roman",
            "Courier": "Courier",
            "Verdana": "helv",            # Using helv as fallback
        }

        # Default style: we store the user's "friendly" name, then map to base-14
        self.text_style = {
            'font': 'Arial',
            'size': 12,
            'color': (0, 0, 0)  # black in RGB (0-1)
        }

        # Create a temporary working copy of the PDF
        try:
            self.temp_dir = tempfile.mkdtemp(prefix="pdf_viewer_")
            temp_filename = os.path.basename(pdf_file_path)
            self.working_pdf_path = os.path.join(self.temp_dir, f"working_{temp_filename}")

            shutil.copy2(pdf_file_path, self.working_pdf_path)
            self.doc = fitz.open(self.working_pdf_path)

            if not self.doc or self.doc.page_count == 0:
                raise ValueError("PDF file is invalid or has no pages.")
        except Exception as e:
            QMessageBox.critical(None, "Error", f"Could not open PDF file:\n{str(e)}")
            self.cleanup_temp_files()
            return

        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(main_layout)

        # Scroll area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        main_layout.addWidget(self.scroll_area)

        # Container for pages
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout()
        self.content_layout.setAlignment(Qt.AlignHCenter)
        self.content_widget.setLayout(self.content_layout)
        self.scroll_area.setWidget(self.content_widget)

        self.page_widgets = []
        self.load_pages()

        # QLineEdit for placing new text
        self.current_text_edit = QLineEdit(self)
        # Make it visible with a semi-transparent background
        self.current_text_edit.setStyleSheet("""
            QLineEdit {
                border: 1px solid #0078d7;
                background: rgba(255, 255, 255, 0.9);
                color: black;
                font-size: 14px;
                padding: 2px;
            }
        """)
        self.current_text_edit.hide()
        self.current_text_edit.returnPressed.connect(self.finish_text_edit)
        
        # Add right-click menu for text editing
        self.current_text_edit.setContextMenuPolicy(Qt.CustomContextMenu)
        self.current_text_edit.customContextMenuRequested.connect(self.show_text_edit_menu)
        # Override focus out to finalize or discard
        orig_focus_out = self.current_text_edit.focusOutEvent

        def custom_focus_out(event):
            typed = self.current_text_edit.text().strip()
            if typed:
                self.finish_text_edit()
            else:
                self.current_text_edit.hide()
            orig_focus_out(event)

        self.current_text_edit.focusOutEvent = custom_focus_out

        # Add zoom buttons at the bottom
        self.zoom_layout = QHBoxLayout()
        self.zoom_in_button = QPushButton("+")
        self.zoom_in_button.clicked.connect(self.zoom_in)
        self.zoom_out_button = QPushButton("-")
        self.zoom_out_button.clicked.connect(self.zoom_out)
        self.zoom_layout.addWidget(self.zoom_out_button)
        self.zoom_layout.addWidget(self.zoom_in_button)
        main_layout.addLayout(self.zoom_layout)

        # Enable text editing mode by default
        self.edit_mode = True
        self.text_placement_mode = True
        
        # Add mouse event handling to each page
        for i, pg_widget in enumerate(self.page_widgets):
            pg_widget.page_label.setMouseTracking(True)
            pg_widget.page_label.mousePressEvent = lambda e, idx=i: self.handle_edit_click(e, idx)
            pg_widget.page_label.setContextMenuPolicy(Qt.CustomContextMenu)
            pg_widget.page_label.customContextMenuRequested.connect(lambda pos, idx=i: self.show_page_context_menu(pos, idx))

        self.dragging = False

        # Start in non-edit mode
        self.edit_mode = False
        self.text_placement_mode = False
        
        # Disable text editing by default
        self.current_text_edit.setEnabled(False)

    def show_text_edit_menu(self, position):
        """Show context menu for text editing."""
        menu = QMenu()
        delete_action = menu.addAction("Delete")
        delete_action.triggered.connect(self.delete_current_text)
        menu.exec_(self.current_text_edit.mapToGlobal(position))

    def delete_current_text(self):
        """Delete the currently selected text."""
        if not hasattr(self.current_text_edit, 'original_word') or not self.current_text_edit.original_word:
            return

        try:
            page = self.doc[self.current_page_index]
            word = self.current_text_edit.original_word
            rect = fitz.Rect(word[0], word[1], word[2], word[3])
            
            # Create and apply redaction
            page.add_redact_annot(rect)
            page.apply_redactions()
            
            self.doc.saveIncr()
            self.show_page(self.current_page_index)
            self.current_text_edit.hide()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not delete text:\n{str(e)}")

    def load_pages(self):
        """Load and render all pages from the PDF doc."""
        for i in range(len(self.doc)):
            page_widget = PageWidget(self.content_widget)
            self.content_layout.addWidget(page_widget)
            self.content_layout.addSpacing(20)

            self.page_widgets.append(page_widget)
            self.show_page(i)

    def show_page(self, index):
        """Render a single PDF page and display it in the PageWidget."""
        if not self.doc:
            return
        try:
            page = self.doc.load_page(index)
            mat = fitz.Matrix(self.zoom, self.zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = QImage(
                pix.samples, pix.width, pix.height,
                pix.stride, QImage.Format_RGB888
            )
            pixmap = QPixmap.fromImage(img)

            page_widget = self.page_widgets[index]
            page_widget.page_label.setPixmap(pixmap)
            page_widget.page_label.adjustSize()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not load page {index+1}:\n{str(e)}")

    def remove_page(self, index):
        """Remove a page from the PDF and update the UI."""
        try:
            page_widget = self.page_widgets.pop(index)
            self.content_layout.removeWidget(page_widget)
            page_widget.deleteLater()

            self.doc.delete_page(index)
            self.doc.saveIncr()

            # Re-render the remaining pages
            for i in range(len(self.page_widgets)):
                self.show_page(i)

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not remove page {index+1}:\n{str(e)}")

    def rotate_all_pages(self, degrees=90):
        """Rotate all pages in the PDF."""
        try:
            for page in self.doc:
                page.set_rotation(degrees)
            self.doc.save(self.working_pdf_path)
            for i in range(len(self.page_widgets)):
                self.show_page(i)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not rotate pages:\n{str(e)}")

    def cleanup_temp_files(self):
        """Clean up temporary files/folders."""
        try:
            if self.doc:
                self.doc.close()
            if hasattr(self, 'temp_dir') and os.path.isdir(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except Exception as ex:
            print(f"Cleanup error: {ex}")

    def closeEvent(self, event):
        self.cleanup_temp_files()
        super().closeEvent(event)

    def save_as(self, new_path):
        """Save working PDF to a new location."""
        try:
            shutil.copy2(self.working_pdf_path, new_path)
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error Saving", f"Could not save:\n{str(e)}")
            return False

    def export_to_excel(self, xlsx_path):
        """Export text from each page into an .xlsx workbook (one sheet per page)."""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Page1"
            for p in range(len(self.doc)):
                if p > 0:
                    ws = wb.create_sheet(title=f"Page{p+1}")
                page = self.doc.load_page(p)
                text = page.get_text("text")
                lines = text.splitlines()
                row = 1
                for line in lines:
                    ws.cell(row=row, column=1, value=line)
                    row += 1
            wb.save(xlsx_path)
            return True
        except Exception as e:
            QMessageBox.warning(self, "Export Failed", f"Excel export failed:\n{str(e)}")
            return False

    # ---------------------
    # Text Editing Methods
    # ---------------------
    def edit_text(self, page_number, old_text, new_text):
        """Search & replace text on a given page. (Simple approach)"""
        if not self.doc:
            return False
        try:
            page = self.doc[page_number]
            matches = page.search_for(old_text)
            if matches:
                # Convert user font name to base-14
                fontname = self.font_map.get(self.text_style['font'], "helv")
                for rect in matches:
                    # Redact out old text
                    page.add_redact_annot(rect)
                    page.apply_redactions()
                    # Insert new text in roughly the same position
                    x0, y0, x1, y1 = rect
                    height = y1 - y0
                    page.insert_text(
                        (x0, y0 + height * 0.8),
                        new_text,
                        fontsize=self.text_style['size'],
                        fontname=fontname,
                        color=self.text_style['color']
                    )
                self.doc.saveIncr()
                self.show_page(page_number)
                return True
            return False
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to edit text:\n{str(e)}")
            return False

    def add_highlight(self, page_number, text):
        """Highlight all occurrences of `text` on a given page."""
        if not self.doc:
            return False
        try:
            page = self.doc[page_number]
            matches = page.search_for(text)
            for rect in matches:
                annot = page.add_highlight_annot(rect)
                annot.update()
            self.doc.saveIncr()
            self.show_page(page_number)
            return True
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to highlight text:\n{str(e)}")
            return False

    def delete_text(self, page_number, text):
        """Delete specific occurrences of `text` on a given page with improved precision."""
        if not self.doc:
            return False
        try:
            page = self.doc[page_number]
            # Get exact word locations with more context
            words = page.get_text("words")
            matches = []
            
            # Find exact matches only
            for word in words:
                if word[4].strip() == text.strip():  # Compare exact text
                    rect = fitz.Rect(word[0], word[1], word[2], word[3])
                    matches.append(rect)
            
            if matches:
                for rect in matches:
                    # Add small padding to avoid overlapping with nearby text
                    rect.x0 -= 1
                    rect.x1 += 1
                    page.add_redact_annot(rect)
                page.apply_redactions()
                self.doc.saveIncr()
                self.show_page(page_number)
                return True
            return False
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to delete text:\n{str(e)}")
            return False

    def find_text(self, page_number, text):
        """Find all occurrences of text on a given page without highlighting."""
        if not self.doc:
            return False
        try:
            page = self.doc[page_number]
            matches = page.search_for(text)
            return len(matches) > 0
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to find text:\n{str(e)}")
            return False




    # ---------------------
    # Adding New Text via Mouse Click
    # ---------------------
    def set_text_style(self, style):
        """Set user's chosen font family, size, color (RGB in 0..1)."""
        self.text_style = style

    def start_edit_mode(self):
        """Enable editing mode."""
        self.edit_mode = True
        self.text_placement_mode = True
        self.setCursor(Qt.IBeamCursor)
        self.current_text_edit.setEnabled(True)
        
        # Enable mouse tracking and click handling for all pages
        for i, pg_widget in enumerate(self.page_widgets):
            pg_widget.page_label.setMouseTracking(True)
            pg_widget.page_label.mousePressEvent = lambda e, idx=i: self.handle_edit_click(e, idx)
        
        # Show status message
        QMessageBox.information(self, "Edit Mode", "PDF editing is now enabled.")

    def stop_edit_mode(self):
        """Disable editing mode."""
        self.edit_mode = False
        self.text_placement_mode = False
        self.setCursor(Qt.ArrowCursor)
        self.current_text_edit.hide()
        self.current_text_edit.setEnabled(False)
        
        # Disable mouse tracking and click handling
        for pg_widget in self.page_widgets:
            pg_widget.page_label.setMouseTracking(False)
            pg_widget.page_label.mousePressEvent = None
        
        # Show status message
        QMessageBox.information(self, "View Mode", "PDF editing is now disabled.")

    def handle_edit_click(self, event, page_idx):
        """Enhanced click handler for text editing."""
        if not self.edit_mode:
            return  # Ignore clicks if not in edit mode
            
        self.current_page_index = page_idx
        pos = event.pos()
        
        # Calculate PDF coordinates with proper scaling
        page = self.doc[page_idx]
        page_rect = page.rect
        scale_x = page_rect.width / self.page_widgets[page_idx].page_label.width()
        scale_y = page_rect.height / self.page_widgets[page_idx].page_label.height()
        
        pdf_x = pos.x() * scale_x
        pdf_y = pos.y() * scale_y

        # Check if there's existing text at click position
        words = page.get_text("words")
        clicked_word = None
        for word in words:
            if (word[0] <= pdf_x <= word[2]) and (word[1] <= pdf_y <= word[3]):
                clicked_word = word
                break

        page_widget = self.page_widgets[page_idx]
        global_pos = page_widget.page_label.mapToGlobal(pos)
        widget_pos = self.mapFromGlobal(global_pos)

        # Position and show text edit
        self.current_text_edit.move(widget_pos)
        if clicked_word:
            # Pre-fill with existing text if clicked on word
            self.current_text_edit.setText(clicked_word[4])
            self.current_text_edit.selectAll()
            self.current_text_edit.original_word = clicked_word
            self.current_text_edit.pdf_position = (clicked_word[0], clicked_word[1])
        else:
            self.current_text_edit.clear()
            self.current_text_edit.original_word = None
            self.current_text_edit.pdf_position = (pdf_x, pdf_y)

        # Set size based on text length or default
        text_width = max(180, len(self.current_text_edit.text()) * 10)
        self.current_text_edit.resize(text_width, 25)
        self.current_text_edit.show()
        self.current_text_edit.setFocus()
        
        # Create context menu
        self.current_text_edit.setContextMenuPolicy(Qt.CustomContextMenu)
        self.current_text_edit.customContextMenuRequested.connect(self.show_text_edit_menu)

    def finish_text_edit(self):
        """Enhanced method to finish text editing."""
        if not self.current_text_edit.isVisible() or self.current_page_index is None:
            return

        new_text = self.current_text_edit.text().strip()
        original_text = self.current_text_edit.original_word[4] if hasattr(self.current_text_edit, 'original_word') and self.current_text_edit.original_word else None

        # If text hasn't changed, just hide the editor without making changes
        if original_text and new_text == original_text:
            self.current_text_edit.hide()
            return

        if not new_text:
            if hasattr(self.current_text_edit, 'original_word') and self.current_text_edit.original_word:
                self.delete_current_text()
            self.current_text_edit.hide()
            return

        try:
            page = self.doc[self.current_page_index]
            
            # If replacing existing text, delete it first
            if hasattr(self.current_text_edit, 'original_word') and self.current_text_edit.original_word:
                word = self.current_text_edit.original_word
                rect = fitz.Rect(word[0], word[1], word[2], word[3])
                page.add_redact_annot(rect)
                page.apply_redactions()

            # Insert new text
            fontname = self.font_map.get(self.text_style['font'], "helv")
            page.insert_text(
                self.current_text_edit.pdf_position,
                new_text,
                fontsize=self.text_style['size'],
                fontname=fontname,
                color=self.text_style['color'],
                render_mode=0  # Ensure text is rendered normally
            )
            
            self.doc.saveIncr()
            self.show_page(self.current_page_index)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not insert text:\n{str(e)}")
        finally:
            self.current_text_edit.hide()

    def zoom_in(self):
        """Zoom in on the current PDF."""
        self.zoom *= 1.2
        for i in range(len(self.page_widgets)):
            self.show_page(i)

    def zoom_out(self):
        """Zoom out on the current PDF."""
        self.zoom /= 1.2
        for i in range(len(self.page_widgets)):
            self.show_page(i)

    def scroll_to_page(self, page_idx):
        """Scroll to ensure the given page is visible."""
        if 0 <= page_idx < len(self.page_widgets):
            widget = self.page_widgets[page_idx]
            self.scroll_area.ensureWidgetVisible(widget)

    def highlight_rect(self, page_idx, rect):
        """Highlight the specified rectangle on the given page."""
        if 0 <= page_idx < len(self.page_widgets):
            page = self.doc[page_idx]
            # Add temporary highlight annotation
            annot = page.add_highlight_annot(rect)
            annot.update()
            self.show_page(page_idx)
            # Store highlight for later removal
            if not hasattr(self, 'temp_highlights'):
                self.temp_highlights = []
            self.temp_highlights.append((page_idx, annot))

    def clear_highlights(self):
        """Clear all temporary highlights."""
        if hasattr(self, 'temp_highlights'):
            for page_idx, annot in self.temp_highlights:
                page = self.doc[page_idx]
                page.delete_annot(annot)
                self.show_page(page_idx)
            self.temp_highlights = []

    def show_page_context_menu(self, pos, page_idx):
        """Show context menu for the PDF page."""
        menu = QMenu()
        
        # Only show edit action if not in edit mode
        if not self.edit_mode:
            edit_action = menu.addAction("Edit")
            edit_action.triggered.connect(lambda: self.start_edit_mode())
        else:
            stop_edit_action = menu.addAction("Stop Editing")
            stop_edit_action.triggered.connect(self.stop_edit_mode)
        
        if menu.actions():  # Only show menu if it has actions
            menu.exec_(self.page_widgets[page_idx].page_label.mapToGlobal(pos))


###############################################################################
# Main Window: Contains menubar, toolbars, and QTabWidget for multiple PDFs
###############################################################################
class PDFViewerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Universal File Viewer & Editor")
        self.resize(1200, 800)

        # Tab Widget
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(True)
        self.tab_widget.tabCloseRequested.connect(self.close_tab)
        self.setCentralWidget(self.tab_widget)

        self.current_color = (0, 0, 0)  # default black
        self.create_menus()
        self.create_toolbars()

        # We'll keep track of whether dark mode is ON or OFF
        self.dark_mode_enabled = False
        self.file_paths = {}  # Store original file paths
        self.setAttribute(Qt.WA_DeleteOnClose, False)  # Prevent main window from closing

    def create_menus(self):
        menubar = self.menuBar()

        # File Menu
        file_menu = menubar.addMenu("File")

        open_action = QAction("Open PDF...", self)
        open_action.triggered.connect(self.open_pdf)
        file_menu.addAction(open_action)

        # Add Save action
        save_action = QAction("Save", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self.save_file)
        file_menu.addAction(save_action)

        # Save As (simplified)
        save_as_action = QAction("Save As", self)
        save_as_action.triggered.connect(self.save_pdf_as)
        file_menu.addAction(save_as_action)

        close_action = QAction("Close Current File", self)
        close_action.triggered.connect(self.close_current_file)
        file_menu.addAction(close_action)

        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # View Menu
        view_menu = menubar.addMenu("View")
        
        # Dark mode action
        dark_mode_action = QAction("Toggle Dark Mode", self)
        dark_mode_action.triggered.connect(self.toggle_dark_mode)
        view_menu.addAction(dark_mode_action)

    def closeEvent(self, event):
        """Handle main window closing."""
        event.accept()

    def create_toolbars(self):
        toolbar = QToolBar("Main Toolbar")
        self.addToolBar(toolbar)

        # Font combo
        self.font_combo = QComboBox()
        # Provide the same keys as our PDFViewWidget.font_map
        self.font_combo.addItems(["Arial", "Times New Roman", "Courier", "Verdana"])
        self.font_combo.setCurrentText("Arial")
        self.font_combo.currentIndexChanged.connect(self.update_text_style)
        toolbar.addWidget(self.font_combo)

        # Font size combo
        self.size_combo = QComboBox()
        sizes = ["8","9","10","11","12","14","16","18","20","22","24","28","32","36","48","72"]
        self.size_combo.addItems(sizes)
        self.size_combo.setCurrentText("12")
        self.size_combo.currentTextChanged.connect(self.update_text_style)
        toolbar.addWidget(self.size_combo)

        # Text color combo
        self.color_combo = QComboBox()
        self.color_combo.addItems(["Black", "Red", "Blue", "Green", "Custom..."])
        self.color_combo.currentTextChanged.connect(self.handle_color_selection)
        toolbar.addWidget(self.color_combo)

    # ---------------------------------------------------------------------
    # FILE MENU ACTIONS
    # ---------------------------------------------------------------------
    def open_pdf(self):
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Open File(s)",
            "",
            "All Supported Files (*.pdf *.xlsx *.xls *.csv *.json *.txt)"
            ";;PDF Files (*.pdf)"
            ";;Excel Files (*.xlsx *.xls)"
            ";;CSV Files (*.csv)"
            ";;JSON Files (*.json)"
            ";;Text Files (*.txt)"
            ";;All Files (*)",
            options=options
        )
        if files:
            for f in files:
                self.add_file_tab(f)

    def add_file_tab(self, file_path):
        try:
            extension = os.path.splitext(file_path)[1].lower()
            viewer = None
            
            if extension == '.pdf':
                viewer = PDFViewWidget(file_path)
            elif extension in ['.xlsx', '.xls']:
                viewer = self.create_excel_viewer(file_path)
            elif extension == '.csv':
                viewer = self.create_csv_viewer(file_path)
            elif extension == '.json':
                viewer = self.create_json_viewer(file_path)
            elif extension == '.txt':
                viewer = self.create_text_viewer(file_path)
            else:
                QMessageBox.warning(self, "Unsupported Format", "This file format is not supported.")
                return

            if viewer:
                filename = os.path.basename(file_path)
                index = self.tab_widget.addTab(viewer, filename)
                self.file_paths[index] = file_path  # Store original file path
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not open file:\n{str(e)}")


    def create_excel_viewer(self, file_path):
        """Create an editable Excel viewer widget with context menu."""
        try:
            # Create table directly without container
            table = QTableWidget()
            df = pd.read_excel(file_path)
            table.setRowCount(df.shape[0])
            table.setColumnCount(df.shape[1])
            table.setHorizontalHeaderLabels(df.columns)

            # Make table editable
            table.setEditTriggers(QTableWidget.DoubleClicked | 
                                QTableWidget.EditKeyPressed |
                                QTableWidget.AnyKeyPressed)

            # Enable selection of entire rows/columns
            table.setSelectionMode(QTableWidget.ExtendedSelection)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            
            # Show row and column headers
            table.horizontalHeader().setVisible(True)
            table.verticalHeader().setVisible(True)
            
            # Make headers clickable
            table.horizontalHeader().setSectionsClickable(True)
            table.verticalHeader().setSectionsClickable(True)

            # Add context menu to headers and cells
            table.setContextMenuPolicy(Qt.CustomContextMenu)
            table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
            table.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
            
            # Connect context menu signals
            table.customContextMenuRequested.connect(
                lambda pos: self.show_table_context_menu(pos, table)
            )
            table.horizontalHeader().customContextMenuRequested.connect(
                lambda pos: self.show_header_context_menu(pos, table, 'column')
            )
            table.verticalHeader().customContextMenuRequested.connect(
                lambda pos: self.show_header_context_menu(pos, table, 'row')
            )

            # Load data
            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    item = QTableWidgetItem(str(df.iloc[row, col]))
                    table.setItem(row, col, item)

            return table
            
        except Exception as e:
            QMessageBox.warning(self, "Excel Error", f"Could not read Excel file:\n{str(e)}")
            return None

    def show_header_context_menu(self, pos, table, header_type):
        """Show context menu for table headers."""
        menu = QMenu()
        
        if header_type == 'column':
            # Get column number from position
            column = table.horizontalHeader().logicalIndexAt(pos)
            if column >= 0:
                action = menu.addAction(f"Delete Column {column + 1}")
                action.triggered.connect(lambda: self.delete_table_columns(table, column, 1))
        else:  # row
            # Get row number from position
            row = table.verticalHeader().logicalIndexAt(pos)
            if row >= 0:
                action = menu.addAction(f"Delete Row {row + 1}")
                action.triggered.connect(lambda: self.delete_table_rows(table, row, 1))
        
        if menu.actions():
            header = table.horizontalHeader() if header_type == 'column' else table.verticalHeader()
            menu.exec_(header.mapToGlobal(pos))

    def show_table_context_menu(self, pos, table):
        """Show context menu for table cells."""
        menu = QMenu()
        
        # Get selected ranges
        ranges = table.selectedRanges()
        if ranges:
            selected_range = ranges[0]
            if selected_range.rowCount() > 0:
                delete_rows = menu.addAction(f"Delete Selected Row(s)")
                delete_rows.triggered.connect(
                    lambda: self.delete_table_rows(table, selected_range.topRow(), selected_range.rowCount())
                )
            
            if selected_range.columnCount() > 0:
                delete_cols = menu.addAction(f"Delete Selected Column(s)")
                delete_cols.triggered.connect(
                    lambda: self.delete_table_columns(table, selected_range.leftColumn(), selected_range.columnCount())
                )
        
        if menu.actions():
            menu.exec_(table.viewport().mapToGlobal(pos))

    def delete_table_rows(self, table, start_row, count):
        """Delete multiple rows from the table."""
        for _ in range(count):
            table.removeRow(start_row)

    def delete_table_columns(self, table, start_col, count):
        """Delete multiple columns from the table."""
        for _ in range(count):
                table.removeColumn(start_col)

    def create_csv_viewer(self, file_path):
        """Create an editable CSV viewer widget with context menu."""
        try:
            # Create table widget
            table = QTableWidget()
            df = pd.read_csv(file_path)
            table.setRowCount(df.shape[0])
            table.setColumnCount(df.shape[1])
            table.setHorizontalHeaderLabels(df.columns)

            # Make table editable
            table.setEditTriggers(QTableWidget.DoubleClicked | 
                                QTableWidget.EditKeyPressed |
                                QTableWidget.AnyKeyPressed)

            # Enable selection of entire rows/columns
            table.setSelectionMode(QTableWidget.ExtendedSelection)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            
            # Show row and column headers
            table.horizontalHeader().setVisible(True)
            table.verticalHeader().setVisible(True)
            
            # Make headers clickable
            table.horizontalHeader().setSectionsClickable(True)
            table.verticalHeader().setSectionsClickable(True)

            # Add context menu to headers and cells
            table.setContextMenuPolicy(Qt.CustomContextMenu)
            table.horizontalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
            table.verticalHeader().setContextMenuPolicy(Qt.CustomContextMenu)
            
            # Connect context menu signals
            table.customContextMenuRequested.connect(
                lambda pos: self.show_table_context_menu(pos, table)
            )
            table.horizontalHeader().customContextMenuRequested.connect(
                lambda pos: self.show_header_context_menu(pos, table, 'column')
            )
            table.verticalHeader().customContextMenuRequested.connect(
                lambda pos: self.show_header_context_menu(pos, table, 'row')
            )

            # Load data
            for row in range(df.shape[0]):
                for col in range(df.shape[1]):
                    item = QTableWidgetItem(str(df.iloc[row, col]))
                    table.setItem(row, col, item)

            return table

        except Exception as e:
            QMessageBox.warning(self, "CSV Error", f"Could not read CSV file:\n{str(e)}")
            return None

    def create_json_viewer(self, file_path):
        """Create an editable JSON viewer that preserves formatting."""
        try:
            with open(file_path, 'r') as f:
                text = f.read()
                data = json.loads(text)
            
            text_edit = QTextEdit()
            text_edit.setPlainText(text)
            
            def on_text_changed():
                try:
                    # Validate JSON while typing
                    json.loads(text_edit.toPlainText())
                    text_edit.setStyleSheet("")
                except json.JSONDecodeError:
                    # Highlight in red if invalid JSON
                    text_edit.setStyleSheet("background-color: #FFE4E1;")
            
            text_edit.textChanged.connect(on_text_changed)
            return text_edit
                
        except Exception as e:
            QMessageBox.warning(self, "JSON Error", f"Could not read JSON file:\n{str(e)}")
            return None

    def create_text_viewer(self, file_path):
        """Create an editable text viewer widget."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
            
            # Use QTextEdit instead of QTextBrowser for editing
            text_edit = QTextEdit()
            text_edit.setPlainText(text)
            
            # Enable editing
            text_edit.setReadOnly(False)
            
            # Add basic text formatting options
            font = QFont("Courier")
            font.setPointSize(10)
            text_edit.setFont(font)
            
            return text_edit
        except Exception as e:
            QMessageBox.warning(self, "Text Error", f"Could not read text file:\n{str(e)}")
            return None

    def close_tab(self, index):
        widget = self.tab_widget.widget(index)
        if widget:
            if hasattr(widget, 'cleanup_temp_files'):
                widget.cleanup_temp_files()
            self.tab_widget.removeTab(index)
        if index in self.file_paths:
            del self.file_paths[index]

    def close_current_file(self):
        idx = self.tab_widget.currentIndex()
        if idx >= 0:
            self.close_tab(idx)

    def save_file(self):
        """Save the current file in its original format."""
        current_widget = self.tab_widget.currentWidget()
        if not current_widget:
            QMessageBox.warning(self, "No File", "No file is open.")
            return

        current_index = self.tab_widget.currentIndex()
        if current_index not in self.file_paths:
            return self.save_pdf_as()

        original_path = self.file_paths[current_index]
        try:
            # Handle different file types
            if isinstance(current_widget, PDFViewWidget):
                success = current_widget.save_as(original_path)
            
            elif isinstance(current_widget, QWidget) and current_widget.layout():
                # Find table widget in container
                table = None
                for i in range(current_widget.layout().count()):
                    item = current_widget.layout().itemAt(i)
                    if item.widget() and isinstance(item.widget(), QTableWidget):
                        table = item.widget()
                        break
                
                if table:
                    # Get data including headers
                    headers = []
                    for col in range(table.columnCount()):
                        header = table.horizontalHeaderItem(col)
                        headers.append(header.text() if header else f"Column {col + 1}")
                    
                    data = []
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            # Handle empty cells
                            cell_value = item.text() if item else ""
                            row_data.append(cell_value)
                        data.append(row_data)
                    
                    # Create DataFrame and save
                    df = pd.DataFrame(data, columns=headers)
                    ext = os.path.splitext(original_path)[1].lower()
                    
                    if ext in ['.xlsx', '.xls']:
                        df.to_excel(original_path, index=False)
                    else:  # CSV
                        df.to_csv(original_path, index=False)
                    success = True
                else:
                    success = False
            
            elif isinstance(current_widget, QTableWidget):
                # Direct table widget
                data = []
                headers = []
                
                for col in range(current_widget.columnCount()):
                    header_item = current_widget.horizontalHeaderItem(col)
                    headers.append(header_item.text() if header_item else f"Column {col+1}")
                
                for row in range(current_widget.rowCount()):
                    row_data = []
                    for col in range(current_widget.columnCount()):
                        item = current_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)
                
                df = pd.DataFrame(data, columns=headers)
                ext = os.path.splitext(original_path)[1].lower()
                if ext in ['.xlsx', '.xls']:
                    df.to_excel(original_path, index=False)
                else:  # CSV
                    df.to_csv(original_path, index=False)
                success = True
            
            elif isinstance(current_widget, (QTextEdit, QTextBrowser)):
                # Save text content directly to file
                with open(original_path, 'w', encoding='utf-8') as f:
                    f.write(current_widget.toPlainText())
                success = True
            
            else:
                success = False

            if success:
                QMessageBox.information(self, "Saved", "File saved successfully!")
            else:
                QMessageBox.warning(self, "Save Failed", "Could not save the file.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not save file:\n{str(e)}")

    def save_pdf_as(self):
        current_widget = self.tab_widget.currentWidget()
        if not current_widget:
            QMessageBox.warning(self, "No File", "Open a file first.")
            return

        # Add back Excel option to filter
        file_filter = (
            "All Supported Files (*.pdf *.xlsx *.xls *.csv *.json *.txt);;"
            "PDF Files (*.pdf);;"
            "Excel Files (*.xlsx *.xls);;"
            "CSV Files (*.csv);;"
            "JSON Files (*.json);;"
            "Text Files (*.txt);;"
            "All Files (*)"
        )

        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, "Save As...", "", file_filter
        )
        
        if not file_path:
            return

        try:
            ext = os.path.splitext(file_path)[1].lower()
            success = False

            # Handle Excel/CSV export
            if ext in ['.xlsx', '.xls', '.csv']:
                # Get the table widget from either direct table or container
                table = None
                if isinstance(current_widget, QWidget) and current_widget.layout():
                    for i in range(current_widget.layout().count()):
                        item = current_widget.layout().itemAt(i)
                        if item.widget() and isinstance(item.widget(), QTableWidget):
                            table = item.widget()
                            break
                elif isinstance(current_widget, QTableWidget):
                    table = current_widget

                if table:
                    data = []
                    headers = []
                    for col in range(table.columnCount()):
                        header_item = table.horizontalHeaderItem(col)
                        headers.append(header_item.text() if header_item else f"Column {col+1}")
                    
                    for row in range(table.rowCount()):
                        row_data = []
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            row_data.append(item.text() if item else "")
                        data.append(row_data)
                    
                    df = pd.DataFrame(data, columns=headers)
                    if ext in ['.xlsx', '.xls']:
                        df.to_excel(file_path, index=False)
                    else:  # CSV
                        df.to_csv(file_path, index=False)
                    success = True

            # Handle PDF export
            elif ext == '.pdf':
                if isinstance(current_widget, PDFViewWidget):
                    success = current_widget.save_as(file_path)
                else:
                    pdf = fitz.open()
                    page = pdf.new_page()
                    if isinstance(current_widget, QTableWidget):
                        text = ""
                        for row in range(current_widget.rowCount()):
                            row_text = []
                            for col in range(current_widget.columnCount()):
                                item = current_widget.item(row, col)
                                row_text.append(item.text() if item else "")
                            text += "\t".join(row_text) + "\n"
                        page.insert_text((50, 50), text)
                    else:
                        page.insert_text((50, 50), current_widget.toPlainText())
                    pdf.save(file_path)
                    success = True

            # Handle JSON/Text
            elif ext in ['.json', '.txt']:
                content = ""
                if isinstance(current_widget, (QTextEdit, QTextBrowser)):
                    content = current_widget.toPlainText()
                elif isinstance(current_widget, QTableWidget):
                    # Convert table to JSON/text
                    data = []
                    for row in range(current_widget.rowCount()):
                        row_data = {}
                        for col in range(current_widget.columnCount()):
                            header = current_widget.horizontalHeaderItem(col)
                            key = header.text() if header else f"Column {col+1}"
                            item = current_widget.item(row, col)
                            row_data[key] = item.text() if item else ""
                        data.append(row_data)
                    if ext == '.json':
                        content = json.dumps(data, indent=2)
                    else:
                        content = str(data)

                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                success = True

            if success:
                QMessageBox.information(self, "Success", "File saved successfully!")
            else:
                QMessageBox.warning(self, "Error", "Could not save in the selected format.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not save file:\n{str(e)}")

    def save_pdf_as_excel(self):
        current_widget = self.tab_widget.currentWidget()
        if not current_widget or not hasattr(current_widget, "export_to_excel"):
            QMessageBox.warning(self, "No PDF", "Open a PDF first.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save As Excel (XLSX)", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        if file_path:
            success = current_widget.export_to_excel(file_path)
            if success:
                QMessageBox.information(self, "Exported", "Exported to Excel successfully!")

    # ---------------------------------------------------------------------
    # VIEW MENU ACTIONS
    # ---------------------------------------------------------------------
    def toggle_dark_mode(self):
        """Toggle dark mode for the entire app."""
        self.dark_mode_enabled = not self.dark_mode_enabled
        if self.dark_mode_enabled:
            self.apply_dark_theme()
        else:
            self.apply_light_theme()

    def apply_dark_theme(self):
        dark_style = """
            QMainWindow {
                background-color: #2b2b2b;
            }
            QWidget {
                color: #dddddd;
                background-color: #3c3c3c;
            }
            QLabel {
                color: #ffffff;
            }
            QLineEdit, QTextEdit, QSpinBox, QComboBox {
                background-color: #2b2b2b;
                color: #ffffff;
                border: 1px solid #5a5a5a;
            }
            QToolBar {
                background-color: #3c3c3c;
            }
            QMenuBar {
                background-color: #3c3c3c;
            }
            QMenuBar::item {
                background: #3c3c3c;
                color: #ffffff;
            }
            QMenu {
                background-color: #3c3c3c;
                color: #ffffff;
            }
            QPushButton {
                background-color: #5a5a5a;
                color: #ffffff;
            }
        """
        self.setStyleSheet(dark_style)

    def apply_light_theme(self):
        self.setStyleSheet("")

    # ---------------------------------------------------------------------
    # TOOLBAR ACTIONS
    # ---------------------------------------------------------------------
    def update_text_style(self):
        current_widget = self.tab_widget.currentWidget()
        if not current_widget:
            return

        font_family = self.font_combo.currentText()
        try:
            size = float(self.size_combo.currentText())
        except ValueError:
            size = 12
        color = self.current_color

        style = {
            'font': font_family,
            'size': size,
            'color': color
        }
        if hasattr(current_widget, "set_text_style"):
            current_widget.set_text_style(style)

    def handle_color_selection(self, color_name):
        color_map = {
            "Black": (0, 0, 0),
            "Red": (1, 0, 0),
            "Blue": (0, 0, 1),
            "Green": (0, 1, 0)
        }
        if color_name == "Custom...":
            c = QColorDialog.getColor()
            if c.isValid():
                self.current_color = (c.red() / 255, c.green() / 255, c.blue() / 255)
                new_label = f"Custom ({c.name()})"
                if self.color_combo.findText(new_label) == -1:
                    self.color_combo.insertItem(0, new_label)
                self.color_combo.setCurrentText(new_label)
                color_map[new_label] = self.current_color
        else:
            self.current_color = color_map.get(color_name, (0, 0, 0))
        self.update_text_style()




def main():
    app = QApplication(sys.argv)
    viewer = PDFViewerApp()
    viewer.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
